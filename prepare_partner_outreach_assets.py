from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_PATH = Path("/workspace/Сколково контакты.xlsx")
MESSAGES_WORKBOOK_PATH = Path("/workspace/Сообщения для outreach.xlsx")
TARGET_SHEETS = ["Екатеринбург", "Казань", "Нижний Новгород", "Краснодарский край"]
OUTPUT_SHEET = "Влад - объединено"
EXCLUDED_RESPONSIBLES = ("аня", "анна", "лена", "елена", "юля")
ADDED_FONT = Font(color="666666")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")


@dataclass(frozen=True)
class AddedInfo:
    email_or_form: str = ""
    phone: str = ""
    social: str = ""
    official_page: str = ""
    confidence: str = ""
    note: str = ""
    sources: str = ""


ADDED_CONTACTS: dict[tuple[str, int], AddedInfo] = {
    ("Екатеринбург", 2): AddedInfo(
        email_or_form="asp@altekproekt.ru",
        phone="+7 (343) 283-07-37; +7 (343) 283-07-30",
        official_page="ООО «Альтек Строй Проект»: altek-stroi-proekt.eb24.ru",
        confidence="средняя-высокая",
        note="Личный TG уже есть в базе. Добавлен корпоративный путь через компанию.",
        sources="https://altek-stroi-proekt.eb24.ru/ ; https://companies.rbc.ru/id/1106674009122-ooo-altek-stroj-proekt/",
    ),
    ("Екатеринбург", 3): AddedInfo(
        email_or_form="expertural@expertural.com",
        phone="+7 (343) 345-03-42",
        official_page="Контакты «Эксперт-Урал»: expert-ural.com/kontakti/ ; страница автора: expert-ural.com/authors/kovalenko-svetlana.html",
        confidence="высокая",
        note="Для первого касания безопаснее идти через редакционный e-mail и ссылку на страницу Светланы.",
        sources="https://expert-ural.com/kontakti/ ; https://expert-ural.com/authors/kovalenko-svetlana.html",
    ),
    ("Екатеринбург", 4): AddedInfo(
        official_page="Официальный сайт бренда: yarmysheva.ru",
        confidence="средняя",
        note="Надежный публичный прямой e-mail не подтвержден; лучше заходить через сайт бренда.",
        sources="https://yarmysheva.ru/",
    ),
    ("Екатеринбург", 5): AddedInfo(
        confidence="низкая",
        note="Без ИНН/точного юрлица надежный публичный контакт не найден; высокий риск ошибки по одноименным компаниям.",
    ),
    ("Екатеринбург", 6): AddedInfo(
        email_or_form="aviharev-vc@yandex.ru",
        phone="+7 (922) 130-44-33; +7 (912) 207-19-54",
        social="VK: vk.com/volunteersekb",
        official_page="Контакты Волонтерского центра",
        confidence="высокая",
        note="Подходит для аккуратного выхода через центр/приемную.",
        sources="https://xn----8sbafbbrdz0awkb4aez0d3f.xn--p1ai/kontakty/ ; https://vk.com/volunteersekb",
    ),
    ("Екатеринбург", 7): AddedInfo(
        phone="+7 (343) 226-02-72",
        official_page="Карточка ООО «КВАРТА»",
        confidence="средняя",
        note="Нужно перепроверить ИНН/юрлицо: в открытых базах встречаются одноименные компании.",
        sources="https://companium.ru/id/1136678003252-kvarta",
    ),
    ("Екатеринбург", 8): AddedInfo(
        email_or_form="fr@lifemart.ru; pr@lifemart.ru",
        phone="+7 912 044-89-64; +7 (800) 700-56-11",
        official_page="Lifemart / франшиза; общий канал сети «Сушкоф»",
        confidence="высокая",
        note="Для партнерства лучше писать через Lifemart fr/pr, а не в лоб на общий контакт сети.",
        sources="https://fr.lifemart.ru/ ; https://www.orgpage.ru/ekaterinburg/sushkof-5973488.html",
    ),
    ("Екатеринбург", 9): AddedInfo(
        email_or_form="Dmitry.Larin@okkam.ru",
        official_page="Профиль команды Okkam и страница контактов",
        confidence="средняя-высокая",
        note="E-mail найден в публичной индексации; перед массовой рассылкой стоит визуально перепроверить страницу.",
        sources="https://okkam.group/team/dmitrij_larin ; https://okkam.group/contacts",
    ),
    ("Екатеринбург", 10): AddedInfo(
        email_or_form="post@bergauf.ru",
        phone="+7 (343) 278-52-94; +7 (343) 310-79-35",
        official_page="Bergauf contacts; региональные контакты Isover/ Saint-Gobain",
        confidence="средняя",
        note="Личного публичного контакта Ирины не найдено; добавлены брендовые каналы компаний.",
        sources="https://bergauf.ru/contacts/ ; https://www.orgpage.ru/ekaterinburg/isover-proizvodstvennaya-1864503.html",
    ),
    ("Екатеринбург", 11): AddedInfo(
        official_page="Карточка ООО ИК «Энергософт» с ИНН и адресом",
        confidence="высокая",
        note="В базе уже есть прямой телефон, TG и сайт; дополнительно подтверждена связка с юрлицом.",
        sources="https://companies.rbc.ru/id/1126671019661-ooo-inzhiniringovaya-kompaniya-energosoft/",
    ),
    ("Екатеринбург", 12): AddedInfo(
        confidence="низкая",
        note="В базе уже есть прямой TG. Дополнительный надежный публичный канал без риска ошибки не найден.",
    ),
    ("Екатеринбург", 13): AddedInfo(
        confidence="низкая",
        note="В базе уже есть прямой TG. Для точного web-enrichment нужен ИНН одной из компаний.",
    ),
    ("Екатеринбург", 14): AddedInfo(
        email_or_form="info@maritol.ru",
        phone="+7 (343) 268-30-61",
        official_page="maritol.ru/about/",
        confidence="высокая",
        note="Добавлен надежный корпоративный контакт компании.",
        sources="https://maritol.ru/about/",
    ),
    ("Екатеринбург", 15): AddedInfo(
        phone="+7 (343) 363-53-25; +7 (922) 020-53-94",
        official_page="autoprocat.ru и карточка 2GIS",
        confidence="средняя",
        note="В базе уже есть прямой TG; e-mail с сайта не подтвердился, поэтому оставлен телефон/сайт.",
        sources="https://autoprocat.ru/ ; https://2gis.ru/firm/70000001038545258",
    ),
    ("Екатеринбург", 16): AddedInfo(
        social="@MyRyadom2020bot",
        official_page="miryadom2020.tilda.ws ; xn--2020-p4d1cbpq3iyb.xn--p1ai",
        confidence="средняя-высокая",
        note="В базе уже есть прямой TG; дополнительно подтверждены официальные страницы проекта.",
        sources="https://miryadom2020.tilda.ws/ ; https://xn--2020-p4d1cbpq3iyb.xn--p1ai/",
    ),
    ("Екатеринбург", 33): AddedInfo(
        phone="+7 (343) 286-11-63",
        official_page="Уралэкспоцентр: uralex.ru/index.php/contacts",
        confidence="низкая",
        note="Только условный корпоративный путь; использовать после ручной проверки, что это именно нужный Александр Баранов.",
        sources="https://www.uralex.ru/index.php/contacts",
    ),
    ("Казань", 2): AddedInfo(
        official_page="Карточка ООО «РЕСУРС» с ИНН/ОГРН и адресом",
        confidence="средняя",
        note="Подтверждена связка с юрлицом, но открытый телефон/e-mail не найден.",
        sources="https://companies.rbc.ru/id/1211600091625-obschestvo-s-ogranichennoj-otvetstvennostyu-resurs/",
    ),
    ("Казань", 6): AddedInfo(
        phone="8 (800) 700-03-88; +7 (843) 250-15-24",
        official_page="marinaizmaylova.ru ; адрес: Казань, ул. Дорожная, 12а",
        confidence="средняя",
        note="Личный e-mail с лендинга не подтвердился; добавлены официальный сайт и корпоративные телефоны из открытых справочников.",
        sources="https://marinaizmaylova.ru/ ; https://mestam.info/ru/kazan/mesto/979036-standartprodmash-dorojnaya-12a",
    ),
    ("Казань", 7): AddedInfo(
        email_or_form="eco@tatar.ru; интернет-приемная",
        phone="+7 (843) 267-68-01; +7 (843) 267-68-02",
        official_page="Контакты Минэкологии РТ",
        confidence="высокая",
        note="Для такого контакта разумнее идти через приемную/официальный канал.",
        sources="https://eco.tatarstan.ru/kontaktnaya-informatsiya.htm ; https://eco.tatarstan.ru/rus/priem.htm",
    ),
    ("Казань", 8): AddedInfo(
        email_or_form="info@romangroup.ru",
        phone="+7 (843) 203-93-87",
        social="VK федерации: vk.com/fctat",
        official_page="ROMAN GROUP contacts; Федерация скалолазания РТ",
        confidence="высокая",
        note="Есть два рабочих пути: через бизнес и через спортивную федерацию.",
        sources="https://romangroup.ru/contacts/ ; https://www.romangroup.ru/company/staff/yakovlev-stanislav-igorevich/ ; https://vk.com/fctat",
    ),
    ("Казань", 10): AddedInfo(
        confidence="низкая",
        note="Без точного юрлица/ИНН не удалось надежно сопоставить контакт.",
    ),
    ("Казань", 11): AddedInfo(
        email_or_form="info@tssenergy.ru",
        phone="+7 (843) 590-36-30",
        official_page="tssenergy.ru",
        confidence="высокая",
        note="Добавлен корпоративный контакт компании; личная привязка требует ручной проверки.",
        sources="http://tssenergy.ru/",
    ),
    ("Казань", 14): AddedInfo(
        email_or_form="tat-chess@mail.ru",
        phone="+7 (843) 236-58-26",
        official_page="Федерация шахмат РТ",
        confidence="средняя",
        note="Добавлен официальный контакт федерации как безопасный вход.",
        sources="https://tat-chess.ru/ ; https://federatsiya-shakhmat-respubliki.orgs.biz/",
    ),
    ("Казань", 16): AddedInfo(
        official_page="Карточка ликвидированного ИП",
        confidence="средняя",
        note="Актуальный публичный контакт не найден; в открытых источниках подтверждается только ликвидация ИП.",
        sources="https://datanewton.ru/contragents/306168515300031",
    ),
    ("Казань", 17): AddedInfo(
        email_or_form="Evadesign@internet.ru",
        official_page="evadesignhome.ru/kontakty",
        confidence="средняя",
        note="В открытых данных найден бренд EVADESIGN; фамилию/юрлицо нужно перепроверить перед отправкой.",
        sources="https://companies.rbc.ru/id/1161690168771-ooo-evo-dizajn/ ; https://evadesignhome.ru/kontakty",
    ),
    ("Казань", 20): AddedInfo(
        email_or_form="tpprt@tpprt.ru",
        phone="+7 (843) 264-62-07",
        official_page="ТПП Татарстана",
        confidence="низкая",
        note="Это скорее ориентир по деловому сообществу, чем подтвержденный прямой контакт Елены Кузнецовой.",
        sources="https://tatarstan.tpprf.ru/ru/contacts/",
    ),
    ("Казань", 21): AddedInfo(
        email_or_form="rgajazov@gmail.com",
        official_page="Карточка ИП/публичный профиль",
        confidence="средняя",
        note="Название клиники в открытых источниках не подтвердилось; использовать аккуратно и лучше сначала в соцсетях.",
        sources="https://companies.rbc.ru/persons/ogrnip/319169000167442-turisticheskaya-kompaniya-7daytravel/ ; https://companium.ru/people/inn/165714096844-gayazov-rustem-rinatovich",
    ),
    ("Казань", 23): AddedInfo(
        email_or_form="ildar.mn82@gmail.com",
        official_page="Карточки ТРИТОН ИНВЕСТ / ПРОГРЕСС ОЙЛ",
        confidence="средняя",
        note="Публичный e-mail найден в открытом досье; перед email-рассылкой лучше перепроверить по дополнительным источникам.",
        sources="https://companies.rbc.ru/id/1211600016319-obschestvo-s-ogranichennoj-otvetstvennostyu-triton-invest/ ; https://companies.rbc.ru/id/1111690086837-ooo-progress-ojl/ ; https://reputation.ru/ogrn/1211600016319",
    ),
    ("Казань", 25): AddedInfo(
        email_or_form="MyCornerMarketing@yandex.ru",
        phone="+7 (800) 600-17-25",
        social="VK: vk.com/mycorner_by_unistroy",
        official_page="mycorner.ru",
        confidence="высокая",
        note="Хороший рабочий путь для партнерского предложения через бренд/маркетинг.",
        sources="https://mycorner.ru/ ; https://realnoevremya.ru/persons/2174-kornilov-anton-yurevich ; https://kazan.spravochnik-rf.ru/stroitelnye-kompanii/949900.html",
    ),
    ("Казань", 29): AddedInfo(
        email_or_form="ABIReception@abdev.ru",
        phone="+7 (843) 205-49-90",
        official_page="akbars-eng.ru",
        confidence="средняя",
        note="В открытых данных есть действующий корпоративный путь через «АК БАРС Инжиниринг», но с вашей пометкой «ликвидирована» нужно сверить, тот ли это человек.",
        sources="https://akbars-eng.ru/ ; https://reputation.ru/inn/166102619176",
    ),
    ("Казань", 33): AddedInfo(
        confidence="средняя",
        note="В базе уже есть TG; дополнительный подтвержденный деловой канал не найден.",
    ),
    ("Казань", 34): AddedInfo(
        official_page="Профиль executive.ru",
        confidence="низкая",
        note="Совпадение с казанским контактом не подтверждено; использовать только после ручной проверки.",
        sources="https://www.e-xecutive.ru/users/1798220-uliya-matunina",
    ),
    ("Казань", 35): AddedInfo(
        official_page="Карточка ИП с совпадающим ФИО",
        confidence="низкая",
        note="Однозначная привязка к нужному человеку не подтверждена.",
        sources="https://companies.rbc.ru/persons/ogrnip/315169000020962-husnutdinov-ajdar-shajhelislamovich/",
    ),
    ("Казань", 37): AddedInfo(
        official_page="imgevent.ru/o-nas",
        confidence="высокая",
        note="В базе уже есть сайт агентства; отдельный публичный e-mail конкретно Виталия быстро не нашелся.",
        sources="https://imgevent.ru/o-nas",
    ),
    ("Казань", 38): AddedInfo(
        official_page="imgevent.ru/o-nas",
        confidence="средняя",
        note="Нужна ручная проверка фамилии/идентификации, но как вход можно использовать общий сайт агентства.",
        sources="https://imgevent.ru/o-nas",
    ),
    ("Нижний Новгород", 2): AddedInfo(
        email_or_form="vita@kis.ru",
        phone="+7 (831) 412-32-17; +7 (910) 393-47-17",
        official_page="vita-print.com",
        confidence="высокая",
        note="Надежный корпоративный путь по компании «Вита-Принт».",
        sources="https://reputation.ru/ogrn/1025203728252 ; https://novgorod.spravker.ru/proizvodstvennyie-predpriyatiya/vita-print.htm ; https://vita-print.com/",
    ),
    ("Нижний Новгород", 6): AddedInfo(
        email_or_form="zakaz@globaltest.ru",
        phone="+7 (83130) 6-77-77 доб. 153",
        official_page="globaltest.ru/kontakty/",
        confidence="высокая",
        note="В базе уже есть мобильный номер; добавлен надежный корпоративный канал для Никиты Козлова.",
        sources="https://www.elec.ru/catalog/globaltest/contacts/ ; https://globaltest.ru/kontakty/",
    ),
    ("Нижний Новгород", 7): AddedInfo(
        phone="8 (800) 500-49-21; +7 (986) 751-15-76",
        official_page="johntruck.ru/contacts/",
        confidence="высокая",
        note="Прямой e-mail не найден; безопаснее идти через контакты/форму John Truck.",
        sources="https://johntruck.ru/contacts/ ; https://www.tbank.ru/business/contractor/legal/1125263004052/",
    ),
    ("Краснодарский край", 2): AddedInfo(
        email_or_form="Форма партнерства на gastreet.com/partners",
        phone="8 (800) 700-93-20",
        official_page="Gastreet partnership page",
        confidence="высокая",
        note="Для такого контакта лучше идти через официальный партнерский вход и уже внутри просить интро на Евгению.",
        sources="https://gastreet.com/partners ; https://2021.gastreet.com/contacts",
    ),
    ("Краснодарский край", 3): AddedInfo(
        social="TG: @Mantera_career",
        official_page="mantera.ru/contacts/ ; блок пресс-службы",
        confidence="средняя",
        note="Личный e-mail Вадима в открытом виде не найден; рабочий путь — пресс-служба/корпоративные контакты группы.",
        sources="https://mantera.ru/contacts/ ; https://mantera.ru/company/ ; https://t.me/Mantera_career",
    ),
    ("Краснодарский край", 6): AddedInfo(
        phone="+7 (961) 587-77-28",
        official_page="Карточка ООО «Альфа-Сервис»",
        confidence="средняя",
        note="Нужно перепроверить ИНН: в открытых базах есть несколько одноименных ООО.",
        sources="https://companies.rbc.ru/id/1122367000733-ooo-alfa-servis/ ; https://sochi.ruspravochnik.com/company/alfa-servis-25/alfa-servis-rossiya-krasnodarskiy-kray-sochi-ulica-yana-fabriciusa-228",
    ),
    ("Краснодарский край", 7): AddedInfo(
        email_or_form="infocenter@kpresort.ru; order@kpresort.ru; sales@kpresort.ru",
        phone="+7 (800) 550-20-20; +7 (862) 245-50-50",
        social="TG: @kp_resort",
        official_page="Контакты курорта «Красная Поляна»",
        confidence="высокая",
        note="Для партнера логичнее писать на infocenter/sales, чем искать персональный адрес.",
        sources="https://ww3.krasnayapolyanaresort.ru/kurort/about/contacts ; https://krasnayapolyanaresort.ru/contact ; https://mantera.ru/company/",
    ),
    ("Краснодарский край", 8): AddedInfo(
        phone="+7 (862) 225-55-35",
        official_page="Карточка ООО «И-ТУР»",
        confidence="средняя-высокая",
        note="Официальный сайт/e-mail не подтвердились, но телефон и связка с юрлицом выглядят рабочими.",
        sources="https://companies.rbc.ru/id/1132366001580-ooo-i-tur/ ; https://sochi.kitabi.ru/firms/i-tur-nesebrskaya-ulica-6",
    ),
    ("Краснодарский край", 9): AddedInfo(
        official_page="alias-group.ru/contacts",
        confidence="высокая",
        note="На странице контактов есть e-mail и TG, но в автоматической выгрузке текст адресов не раскрылся; лучше открыть страницу вручную перед отправкой.",
        sources="https://alias-group.ru/contacts ; https://alias-group.ru/ ; https://companium.ru/id/1212300024640-ehlias-grupp",
    ),
    ("Краснодарский край", 11): AddedInfo(
        confidence="низкая",
        note="Без компании/ИНН надежный публичный контакт не найден.",
    ),
    ("Краснодарский край", 16): AddedInfo(
        email_or_form="bsfc@bsfc.com",
        phone="8 (800) 700-1-800",
        official_page="bsfc.com/about/contacts/",
        confidence="высокая",
        note="Хороший официальный вход через агентство/группу.",
        sources="https://bsfc.com/about/contacts/",
    ),
    ("Краснодарский край", 17): AddedInfo(
        official_page="Карточка ООО «Ванвин» с адресом",
        confidence="высокая",
        note="Компания подтверждена, но публичный e-mail/телефон не найден.",
        sources="https://companies.rbc.ru/id/1202300013145-obschestvo-s-ogranichennoj-otvetstvennostyu-vanvin/",
    ),
    ("Краснодарский край", 21): AddedInfo(
        email_or_form="Форма заявки на itr2050.ru",
        official_page="itr2050.ru/komanda.html ; itr2050.ru",
        confidence="высокая",
        note="В базе уже есть сайт; дополнительно отмечен рабочий вход через форму заявки.",
        sources="https://itr2050.ru/komanda.html ; https://itr2050.ru/",
    ),
    ("Краснодарский край", 28): AddedInfo(
        email_or_form="mischenkoi@develug.ru",
        phone="+7 861 279-45-67; +7 861 279-46-32",
        official_page="Карточки ООО СК «Екатеринодар-Сити»",
        confidence="средняя",
        note="Контакты взяты из открытых досье/агрегаторов; лучше перепроверить на официальных документах перед email-рассылкой.",
        sources="https://zachestnyibiznes.ru/company/ul/1092310003389_2310140354_OOO-SK-EKATERINODAR-SITI ; https://companies.rbc.ru/id/1092310003389-ooo-stroitelnaya-kompaniya-ekaterinodar-siti/",
    ),
}


# Второй проход: дополняем таблицу более глубоким поиском по официальным сайтам,
# карточкам компаний и публичным профилям.
ADDED_CONTACTS.update(
    {
        ("Екатеринбург", 12): AddedInfo(
            email_or_form="lgst@tmr.rest; mr@tmr.rest",
            phone="+7 965 544-65-11; +7 (343) 301-02-22",
            official_page="shrmps-brgrs.ru/contact ; comunale.rest/about ; morskaya10.rest/contacts",
            confidence="высокая",
            note="Публичные рабочие контакты ресторанных проектов, связанных с Андреем Максимовым; это каналы бизнеса, не личный адрес.",
            sources="https://shrmps-brgrs.ru/contact ; https://comunale.rest/about ; https://morskaya10.rest/contacts",
        ),
        ("Екатеринбург", 13): AddedInfo(
            official_page="ООО «Бипоинт» и ООО «Бипоинт Логистик»: публичные карточки компаний с адресами",
            confidence="средняя",
            note="Прямые телефон/email в открытом доступе не подтвердились, но добавлены надежные карточки юрлиц и адреса для верификации.",
            sources="https://www.tbank.ru/business/contractor/legal/1156658104382/ ; https://www.tbank.ru/business/contractor/legal/1146670016052/",
        ),
        ("Екатеринбург", 14): AddedInfo(
            email_or_form="info@maritol.ru; snitsar@maritol.ru",
            phone="+7 (343) 268-30-61",
            official_page="maritol.ru/about/ ; maritol.ru/contact/",
            confidence="высокая",
            note="Добавлен не только общий, но и публичный корпоративный e-mail Семена Сницара.",
            sources="https://maritol.ru/about/ ; https://maritol.ru/contact/",
        ),
        ("Екатеринбург", 17): AddedInfo(
            official_page="Карточка ООО «Мастер Трак»: Екатеринбург, ул. Кирова, 40б, пом. 272",
            confidence="средняя",
            note="Прямой телефон/email не раскрыты, но подтверждены компания и адрес.",
            sources="https://companies.rbc.ru/id/1216600016709-obschestvo-s-ogranichennoj-otvetstvennostyu-master-trak/",
        ),
        ("Екатеринбург", 19): AddedInfo(
            email_or_form="ada@ak24.ru",
            official_page="Карточка ООО «Копир Плюс»: Екатеринбург, ул. Мартовская, 6, пом. 301",
            confidence="средняя",
            note="Публичный e-mail найден в открытой карточке компании; по ИТ-Альянсу надежная публичная привязка не подтвердилась.",
            sources="https://www.tbank.ru/business/contractor/legal/1169658109763/ ; https://reputation.ru/ogrn/1169658109763",
        ),
        ("Екатеринбург", 20): AddedInfo(
            official_page="Личный сайт rvaliev.com ; сайт компании npcat.ru",
            confidence="высокая",
            note="В базе уже есть прямые телефон и TG; дополнительно подтверждены публичный сайт Рафаила и сайт компании.",
            sources="https://rvaliev.com/ ; https://npcat.ru/",
        ),
        ("Екатеринбург", 21): AddedInfo(
            email_or_form="im@moveli.ru; afimoll@moveli.ru; aviapark@moveli.ru",
            phone="+7 (916) 503-99-93",
            social="TG: @moveli_ru, @moveli_shop; VK: vk.com/moveli_store; WhatsApp: wa.me/79165039993",
            official_page="moveli.ru/page/about-us ; moveli.ru/page/contacts",
            confidence="высокая",
            note="У бренда опубликованы сразу несколько рабочих e-mail магазинов/направлений; для первого касания лучше использовать общий канал + TG бренда.",
            sources="https://www.moveli.ru/page/about-us ; https://www.moveli.ru/page/contacts",
        ),
        ("Екатеринбург", 23): AddedInfo(
            email_or_form="info@elittent.org; e.v@elittent.org",
            phone="+7 (343) 382-80-68; +7 (922) 152-55-25; +7 (922) 110-07-33",
            social="VK: vk.com/elittent",
            official_page="elittent.org/company/ ; elittent.org/company/requisites/",
            confidence="высокая",
            note="Сильный корпоративный набор: общий e-mail, персональный корпоративный e-mail, телефоны и VK.",
            sources="https://elittent.org/company/ ; https://elittent.org/company/requisites/",
        ),
        ("Екатеринбург", 24): AddedInfo(
            email_or_form="lgst@tmr.rest",
            phone="+7 (343) 305-05-46; +7 (343) 305-05-40",
            official_page="rtt-96.ru/kontakty ; карточка ООО «Новый Вкус»",
            confidence="высокая",
            note="Добавлены официальные контакты ресторанной группы/компании, связанной с Сергеем Тихненко.",
            sources="https://www.rtt-96.ru/kontakty ; https://companies.rbc.ru/id/1086671015166-ooo-novyij-vkus/",
        ),
        ("Екатеринбург", 25): AddedInfo(
            email_or_form="sakurapark.info@mail.ru",
            phone="+7 932 609-88-71",
            official_page="Официальная страница контактов Sakura Park",
            confidence="высокая",
            note="На странице прямо указан директор Евгения Сизганова.",
            sources="https://kp96.ru/contacts/",
        ),
        ("Екатеринбург", 26): AddedInfo(
            phone="8 (953) 04-04-045",
            official_page="izi.show/contacts/ ; офис: Екатеринбург, ул. Энгельса 36, этаж 7, оф. 714/1",
            confidence="высокая",
            note="Добавлен официальный телефон и страница контактов IZI.show; e-mail публично не раскрыт.",
            sources="https://izi.show/contacts/ ; https://izi.show/about/",
        ),
        ("Екатеринбург", 27): AddedInfo(
            official_page="Карточка ООО «ИМ Диджитал»: Екатеринбург, ул. Ткачей, 23, офис 708",
            confidence="средняя",
            note="Подтверждено юрлицо и адрес; прямой рабочий телефон/email не раскрыт.",
            sources="https://companies.rbc.ru/id/1226600006038-obschestvo-s-ogranichennoj-otvetstvennostyu-im-didzhital/",
        ),
        ("Екатеринбург", 28): AddedInfo(
            email_or_form="920618@mail.ru",
            social="Instagram: instagram.com/gareevrr",
            official_page="Личный сайт gareev.com",
            confidence="высокая",
            note="На личном сайте Романа Гареева опубликован e-mail; TG уже был в базе.",
            sources="https://gareev.com/",
        ),
        ("Екатеринбург", 31): AddedInfo(
            phone="+7 (343) 373-77-86",
            official_page="elisey-mag.ru / карточка сети «Елисей»",
            confidence="средняя",
            note="Это, скорее, коммутатор/контакт сети, а не прямой номер Александра Оглоблина.",
            sources="https://www.orgpage.ru/ekaterinburg/elisey-set-supermarketov-2021789.html",
        ),
        ("Екатеринбург", 32): AddedInfo(
            phone="+7 (343) 283-08-80",
            social="TG: @goodcomdeliveryiset, @goodcommunitydelivery",
            official_page="goodcom.rest/breadway ; dostavka.goodcom.rest/contacts/",
            confidence="высокая",
            note="Публичные бизнес-контакты ресторанного холдинга; это рабочий вход на группу, не личный контакт Руслана.",
            sources="https://goodcom.rest/breadway ; https://dostavka.goodcom.rest/contacts/ ; https://goodcom.rest/",
        ),
        ("Екатеринбург", 34): AddedInfo(
            phone="+7 (343) 226-06-62",
            social="TG: @initki, @initki_knit; VK: vk.com/initki_96",
            official_page="get.initki.ru/about_school",
            confidence="высокая",
            note="Официальные контакты школы/проекта iNitki.",
            sources="https://get.initki.ru ; https://get.initki.ru/about_school",
        ),
        ("Екатеринбург", 36): AddedInfo(
            email_or_form="zakaz@grata-mebel.ru; partner@grata-mebel.ru; director@grata-mebel.ru",
            phone="+7 (982) 622-98-88; 8 (804) 333-14-22",
            official_page="grata-mebel.ru/kontakty",
            confidence="высокая",
            note="У компании опубликовано несколько сильных e-mail по направлениям; для первого касания безопаснее писать на partner/zakaz.",
            sources="https://grata-mebel.ru ; https://grata-mebel.ru/kontakty",
        ),
        ("Екатеринбург", 37): AddedInfo(
            email_or_form="office@ema.su",
            phone="8 (343) 300-40-11",
            official_page="emaholding.ru/#contacts",
            confidence="высокая",
            note="Официальный корпоративный контакт холдинга ЭМА.",
            sources="https://emaholding.ru",
        ),
        ("Екатеринбург", 38): AddedInfo(
            email_or_form="info@19agency84.ru; newbusiness@19agency84.com",
            social="TG: @aloha19agency84; VK: vk.com/19agency84",
            official_page="19agency84.ru/contacts",
            confidence="высокая",
            note="Для партнёрского касания особенно полезен newbusiness@19agency84.com.",
            sources="https://19agency84.ru ; https://19agency84.ru/contacts",
        ),
        ("Екатеринбург", 40): AddedInfo(
            email_or_form="p-service@p-service.pro",
            phone="+7 (343) 266-38-19",
            official_page="p-service.pro/about/",
            confidence="высокая",
            note="Подтвержден корпоративный e-mail и телефон компании.",
            sources="https://p-service.pro ; https://p-service.pro/about/",
        ),
        ("Екатеринбург", 41): AddedInfo(
            official_page="Карточки ООО «ФРЕЙМ» и ООО «МАН ГРУПП» с адресами и ролями",
            confidence="высокая",
            note="Добавлены надежные карточки юрлиц и адреса, хотя публичные телефон/email на страницах не раскрыты.",
            sources="https://companies.rbc.ru/id/1116604000842-ooo-man-grupp/ ; https://companies.rbc.ru/id/1146678013569-ooo-man-invest/",
        ),
        ("Екатеринбург", 42): AddedInfo(
            official_page="Карточки ООО «Девайс», ООО «Новая Магистраль», ООО «ГРИТ» с ролями и адресами",
            confidence="высокая",
            note="Публичные карточки подтверждают роль/компании и помогают верифицировать контакт; прямой телефон уже есть в базе.",
            sources="https://companies.rbc.ru/id/1126681001292-ooo-dejvis/ ; https://companies.rbc.ru/id/1126681001930-ooo-novaya-magistral/ ; https://companies.rbc.ru/id/1226600056517-ooo-grit/",
        ),
        ("Екатеринбург", 43): AddedInfo(
            email_or_form="info@micro-climate.com",
            phone="8 (800) 600-71-98",
            social="TG: @Microklimattechpod_bot, @microclimate_industry; VK: vk.com/micro_climate",
            official_page="micro-climate.com/about ; профиль автора на сайте",
            confidence="высокая",
            note="Собран сильный набор корпоративных каналов производителя.",
            sources="https://micro-climate.com/blog/avtory-bloga/konstantin-petrov ; https://micro-climate.com/about",
        ),
        ("Казань", 6): AddedInfo(
            email_or_form="marinaizmaylova@yandex.ru",
            phone="+7 (919) 622-48-68",
            social="Instagram: instagram.com/marina.izmailova",
            official_page="marinaizmaylova.ru",
            confidence="высокая",
            note="Удалось найти уже не только корпоративный, но и персональный публичный e-mail Марины.",
            sources="https://marinaizmaylova.ru/",
        ),
        ("Казань", 8): AddedInfo(
            email_or_form="info@romangroup.ru; info@fctat.ru",
            phone="+7 (843) 203-93-87; +7 996 644-94-36",
            social="VK: vk.com/fctat",
            official_page="ROMAN GROUP contacts; страница Яковлева на сайте федерации",
            confidence="высокая",
            note="Добавлен второй рабочий e-mail и публичный телефон с профильной страницы.",
            sources="https://romangroup.ru/contacts/ ; https://fctat.ru/federation/sotrudniki/yakovlev-stanislav-igorevich/ ; https://vk.com/fctat",
        ),
        ("Казань", 10): AddedInfo(
            official_page="Карточка ИП с совпадающим ФИО/ИНН/ОГРНИП",
            confidence="низкая",
            note="Нашёлся действующий ИП Исмагилова Ильнара, но он плохо бьётся с вашей пометкой про ликвидированную авто-компанию; использовать только после ручной проверки.",
            sources="https://check.tochka.com/company/317169000070088/",
        ),
        ("Казань", 33): AddedInfo(
            official_page="Карточка ИП Минниханова Азата Раисовича с адресом и реквизитами",
            confidence="средняя",
            note="Прямых телефона/email не найдено; TG уже есть в базе. Добавлены реквизиты для верификации контакта.",
            sources="https://www.tenderer.ru/contragent/163502515514-316169000104697-individualnyy-predprinimatel-minnihanov-azat-raisovich-sabinskiy-r-n ; https://t.me/azatm016",
        ),
        ("Казань", 37): AddedInfo(
            email_or_form="hello@imgevent.ru",
            phone="+7 (962) 572-13-10; 8 (800) 700-51-40",
            social="TG: @IMG_Event; WhatsApp: wa.me/+79625721310",
            official_page="imgevent.ru/contacts",
            confidence="высокая",
            note="Для первого касания лучше использовать hello@imgevent.ru или TG агентства.",
            sources="https://imgevent.ru/o-nas ; https://imgevent.ru/contacts",
        ),
        ("Казань", 38): AddedInfo(
            email_or_form="hello@imgevent.ru",
            phone="+7 (962) 572-13-10; 8 (800) 700-51-40",
            social="TG: @IMG_Event; WhatsApp: wa.me/+79625721310",
            official_page="imgevent.ru/contacts",
            confidence="средняя-высокая",
            note="Используйте как рабочий вход через агентство; фамилия в исходнике требует ручной проверки.",
            sources="https://imgevent.ru/o-nas ; https://imgevent.ru/contacts",
        ),
        ("Краснодарский край", 2): AddedInfo(
            email_or_form="info@gastreet.com; pr@gastreet.com; форма партнерства на gastreet.com/partners",
            phone="8 (800) 700-93-20; +7 967 696-99-20; +7 928 667-07-06",
            official_page="gastreet.com/contacts ; gastreet.com/partners",
            confidence="высокая",
            note="Добавлены уже не только партнерская форма, но и общий/PR e-mail и телефоны Gastreet.",
            sources="https://gastreet.com/contacts ; https://gastreet.com/partners ; https://2023.gastreet.com/contacts",
        ),
        ("Краснодарский край", 3): AddedInfo(
            email_or_form="info@mantera-group.com; globalsales@mantera-group.com; pr@mantera-group.com",
            phone="8 (800) 100-78-62",
            social="TG: @Mantera_career, @mantera",
            official_page="mantera.ru/company/ ; mantera.ru/contacts/",
            confidence="высокая",
            note="Теперь есть не только TG, но и несколько официальных e-mail группы MANTERA.",
            sources="https://mantera.ru/company/ ; https://mantera.ru/contacts/",
        ),
        ("Краснодарский край", 9): AddedInfo(
            email_or_form="info@alias-group.ru",
            phone="+7 (861) 205-09-84",
            official_page="alias-group.ru/contacts",
            confidence="высокая",
            note="Подтвержден общий официальный e-mail и телефон Alias Group.",
            sources="https://alias-group.ru/contacts ; https://alias-group.ru/ ; https://companium.ru/id/1212300024640-ehlias-grupp",
        ),
        ("Краснодарский край", 11): AddedInfo(
            official_page="Карточки ИП Гореловой Анны Александровны и связанных компаний",
            confidence="средняя",
            note="По ФИО+региону нашлась Анна Горелова из Краснодара с ИП и девелоперскими активами, но сфера расходится с исходной строкой — перед использованием нужна ручная проверка.",
            sources="https://companies.rbc.ru/persons/ogrnip/321237500218354-gorelova-anna-aleksandrovna/ ; https://companies.rbc.ru/persons/inn/231004660011/",
        ),
        ("Краснодарский край", 21): AddedInfo(
            email_or_form="itr@itr2050.ru",
            phone="+7 (999) 217-57-72",
            official_page="itr2050.ru ; itr2050.ru/komanda.html",
            confidence="высокая",
            note="Удалось найти прямой e-mail и публичный телефон ИТР, что лучше формы заявки.",
            sources="https://itr2050.ru/ ; https://itr2050.ru/komanda.html",
        ),
    }
)


SOCIAL_MESSAGES = [
    (
        "Самый мягкий заход",
        "Сообщение 1",
        "Здравствуйте, {{Имя}}!\nМеня зовут Влад, я из фонда «БольшеЧемМожешь». Пишу очень аккуратно: возможно, мы раньше уже пересекались по спортивным/партнерским темам, а возможно нет.\nПодскажите, можно буквально в 2 сообщениях рассказать, зачем написал?",
        "Холодная база; нужно сначала получить микро-согласие на диалог.",
        "Получить разрешение продолжить переписку.",
    ),
    (
        "Самый мягкий заход",
        "Сообщение 2",
        "Спасибо! Мы вместе со СКОЛКОВО готовим «Сколковскую милю» — благотворительный забег 20 июня 2026 на кампусе Школы в поддержку наших подопечных с ДЦП и другими двигательными нарушениями.\nСейчас ищу 2-3 партнеров, кому может быть интересна не просто благотворительность, а понятная интеграция в событие: участие в площадке, подарки участникам, брендинг, упоминания в материалах СКОЛКОВО и фонда.\nЕсли тема вам в целом близка, могу отправить очень короткую презентацию или сразу написать 3 варианта участия без простыни.",
        "После любого нейтрального или положительного ответа.",
        "Получить согласие на отправку презентации или короткой выжимки.",
    ),
    (
        "Самый мягкий заход",
        "Сообщение 3",
        "Если вам неудобно это смотреть лично, буду благодарен, если подскажете, кто у вас обычно смотрит такие партнерства: маркетинг, PR, ESG/CSR, HR или спецпроекты. Тогда я напишу уже точечно и без лишних касаний.",
        "Если человек сам не ведет вопрос.",
        "Получить ЛПР или нужного коллегу.",
    ),
    (
        "Быстрее к сути",
        "Сообщение 1",
        "{{Имя}}, здравствуйте!\nЯ Влад, фонд «БольшеЧемМожешь». Ищу партнера на «Сколковскую милю 2026» — благотворительный забег на кампусе МШУ СКОЛКОВО 20 июня.\nПодумал о вас, потому что у вас сильная связка с {{компания / индустрия}}. Могу прислать 1 короткий файл и 3 возможных формата участия?",
        "Когда есть хорошее попадание по бизнесу или индустрии.",
        "Получить согласие на короткий файл.",
    ),
    (
        "Быстрее к сути",
        "Сообщение 2",
        "Коротко по сути:\n- 300+ участников и сообщество СКОЛКОВО;\n- событие открывает День выпускника;\n- интеграции от подарков/сертификатов до партнерского пакета;\n- это поддержка фонда, который развивает инклюзивный спорт и выводит на старты людей с тяжелыми двигательными нарушениями.\nЕсли интересно, пришлю презентацию и предложу вариант именно под ваш формат, без абстрактного «давайте поддержите».",
        "Если собеседник готов читать по делу.",
        "Вывести на просмотр презентации.",
    ),
    (
        "Статусный контакт",
        "Сообщение 1",
        "Здравствуйте, {{Имя}}.\nМеня зовут Влад, я представляю фонд «БольшеЧемМожешь». Понимаю, что пишу без предварительного знакомства, поэтому коротко и по делу.\nСовместно со СКОЛКОВО готовим благотворительный забег «Сколковская миля» на 20 июня 2026. Ищу несколько точных партнерств, где социальная задача сочетается с сильной деловой аудиторией и аккуратной интеграцией бренда.\nЕсли допустимо, направлю краткую презентацию. Если удобнее, подскажите, пожалуйста, коллегу, который смотрит такие вопросы.",
        "Для статусных людей, руководителей, министерств, крупных холдингов.",
        "Получить допуск или контакт коллеги.",
    ),
    (
        "Если контакт может не помнить контекст",
        "Сообщение 1",
        "Здравствуйте, {{Имя}}!\nЯ Влад, фонд «БольшеЧемМожешь». Возможно, мы не знакомы лично и мой контакт всплыл без контекста, поэтому сначала коротко объясню, зачем пишу.\nМы вместе со СКОЛКОВО собираем партнеров на благотворительный забег «Сколковская миля 2026». Это не массовый спам по базе: я пишу точечно людям и компаниям, кому может подойти формат партнерства вокруг спорта, сообщества и социальной повестки.\nМожно отправлю короткое описание, а вы уже решите, стоит ли продолжать разговор?",
        "Когда боитесь, что контакт воспримет сообщение как спам.",
        "Снизить защиту и получить ответ.",
    ),
    (
        "Через просьбу о совете",
        "Сообщение 1",
        "{{Имя}}, добрый день!\nНужен ваш короткий совет как человека, который хорошо понимает {{индустрия / рынок}}.\nМы с фондом «БольшеЧемМожешь» и СКОЛКОВО готовим благотворительный забег в июне 2026 и ищем уместные партнерства.\nМожно в двух сообщениях опишу механику, а вы честно скажете, это вообще может быть релевантно вам / вашей компании или лучше идти к другому профилю?",
        "Когда нужен максимально мягкий вход на холодную базу.",
        "Получить диалог без ощущения продажи в лоб.",
    ),
]

FOLLOW_UP_MESSAGES = [
    (
        "Follow-up 1",
        "Через 3-5 дней",
        "{{Имя}}, аккуратно возвращаюсь к сообщению ниже.\nЕсли тема партнерства для «Сколковской мили» в принципе может быть вам релевантна, я пришлю совсем короткую выжимку на 1 минуту чтения.\nЕсли нет, тоже ок — просто сориентируйте, чтобы я вас не беспокоил.",
    ),
    (
        "Follow-up 2",
        "Через 5-7 дней после первого follow-up",
        "Здравствуйте! Оставлю последнее короткое касание.\nИщу контакт внутри вашей команды, кому корректно показать партнерское предложение по «Сколковской миле 2026». Если подскажете такого человека, дальше уже адресно напишу ему.",
    ),
]

EMAIL_MESSAGES = [
    (
        "Полное письмо",
        "Партнерство на «Сколковскую милю 2026» совместно со СКОЛКОВО",
        "Здравствуйте, {{Имя}}!\n\nМеня зовут Влад, я представляю благотворительный фонд «БольшеЧемМожешь».\n\nПишу вам по поводу «Сколковской мили 2026» — благотворительного забега, который мы проводим совместно с МШУ СКОЛКОВО 20 июня 2026 на кампусе Школы. Забег открывает программу Дня выпускника и проходит в поддержку подростков и взрослых с ДЦП и другими двигательными нарушениями — подопечных нашего фонда.\n\nПочему решил написать именно вам: мне кажется, что для {{Компания}} здесь может быть не только благотворительная составляющая, но и осмысленное партнерство с понятной интеграцией:\n- присутствие бренда на площадке и в материалах события;\n- участие через продукт, сервис, подарки или спецформат;\n- доступ к деловому сообществу СКОЛКОВО и аудитории участников;\n- социальный проект с понятным эффектом и сильной репутационной рамкой.\n\nСейчас рассматриваем несколько форматов участия — от продуктовой поддержки и призов до партнерских пакетов 100 000 / 300 000 / 500 000 / 800 000 ₽.\n\nЕсли тема вам в целом откликается, я отправлю короткую презентацию и предложу 2-3 варианта сотрудничества именно под ваш профиль.\n\nЕсли такие вопросы у вас ведет другой коллега, буду благодарен, если подскажете, кому корректнее написать.\n\nС уважением,\nВладислав\nФонд «БольшеЧемМожешь»\n{{телефон}}\n{{email}}",
        "Когда есть шанс, что письмо реально прочитают и нужен полный контекст.",
    ),
    (
        "Короткое письмо",
        "Кому у вас корректно показать предложение по «Сколковской миле 2026»?",
        "Здравствуйте, {{Имя}}!\n\nЯ Владислав из фонда «БольшеЧемМожешь». Мы вместе со СКОЛКОВО ищем партнеров на благотворительный забег «Сколковская миля 2026» (20 июня, кампус МШУ СКОЛКОВО).\n\nПодумал, что это может быть потенциально релевантно {{Компания}}: здесь можно зайти либо через полноценное партнерство, либо через продукт/призы/интеграцию на площадке.\n\nЕсли можно, пришлю короткую презентацию.\nЕсли это не к вам — подскажите, пожалуйста, кому корректнее написать внутри команды.\n\nС уважением,\nВладислав\nФонд «БольшеЧемМожешь»",
        "Для холодного входа на корпоративный e-mail.",
    ),
]

PROCESS_RECOMMENDATIONS = [
    ("Принцип", "Первое касание должно открывать диалог, а не продавать."),
    ("CTA", "Лучший первый CTA: не «давайте сразу 500 000», а «кому корректно показать 1-страничное предложение / презентацию»."),
    ("Приоритет", "Сначала пишите контактам с прямыми TG/e-mail и высокой или средней уверенностью."),
    ("Волны", "Разделите базу на A/B/C/D: прямой контакт / корпоративный канал / только сайт-форма / нужна ручная верификация."),
    ("Воронка", "Добавьте в таблицу статусы: не писал / 1 касание / follow-up 1 / есть диалог / попросили материалы / передан ЛПР / неактуально."),
    ("Статусные контакты", "Для министерств, холдингов и крупных девелоперов лучше заходить с просьбой дать коллегу по партнерствам / PR / CSR."),
]


STANDARD_COLUMNS = [
    "Источник",
    "ПРАКТИКУМ",
    "ГОРОД",
    "ФИО",
    "Номер телефона ",
    "Телеграм ",
    "Ответственный а контакт/ кто связывется ",
    "ИНДУСТРИЯ",
    "КОМПАНИЯ",
    "САЙТ",
    "Добавлено: email / форма",
    "Добавлено: телефон",
    "Добавлено: соцсеть / канал",
    "Добавлено: оф. контакты / страница",
    "Добавлено: уверенность",
    "Добавлено: комментарий",
    "Добавлено: источники",
]


def cell_value(ws, row: int, header_map: dict[str, int], token: str) -> str | None:
    for header, column in header_map.items():
        if token in header:
            return ws.cell(row, column).value
    return None


def normalize_header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            mapping[str(value).strip()] = col
    return mapping


def should_skip_responsible(value: object) -> bool:
    text = str(value or "").lower()
    return any(name in text for name in EXCLUDED_RESPONSIBLES)


def default_added_info(phone: object, tg: object, site: object) -> AddedInfo:
    has_direct = any(
        str(value).strip() and str(value).strip().lower() != "нет"
        for value in (phone, tg, site)
    )
    if has_direct:
        return AddedInfo(
            confidence="низкая",
            note="В исходной строке уже есть прямой канал; надежного дополнительного публичного контакта без риска ошибки не найдено.",
        )
    return AddedInfo(
        confidence="низкая",
        note="Надежный публичный контакт не найден; нужна ручная проверка ИНН/юрлица или ассистента компании.",
    )


def add_row(ws_out, values: list[object], added_info: AddedInfo) -> None:
    row_index = ws_out.max_row + 1
    row_values = values + [
        added_info.email_or_form,
        added_info.phone,
        added_info.social,
        added_info.official_page,
        added_info.confidence,
        added_info.note,
        added_info.sources,
    ]
    for col_idx, value in enumerate(row_values, start=1):
        cell = ws_out.cell(row=row_index, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx > 10 and value:
            cell.font = ADDED_FONT


def autosize(ws) -> None:
    widths = {
        "A": 18,
        "B": 12,
        "C": 18,
        "D": 28,
        "E": 20,
        "F": 22,
        "G": 18,
        "H": 18,
        "I": 52,
        "J": 30,
        "K": 30,
        "L": 22,
        "M": 24,
        "N": 34,
        "O": 14,
        "P": 46,
        "Q": 52,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_messages_workbook() -> None:
    wb = Workbook()
    ws_social = wb.active
    ws_social.title = "Соцсети"
    ws_follow = wb.create_sheet("Follow-up")
    ws_email = wb.create_sheet("Email")
    ws_process = wb.create_sheet("Рекомендации")

    social_headers = ["Сценарий", "Этап", "Текст", "Когда использовать", "Цель / CTA"]
    for col, header in enumerate(social_headers, start=1):
        cell = ws_social.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(SOCIAL_MESSAGES, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_social.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    follow_headers = ["Сообщение", "Когда отправлять", "Текст"]
    for col, header in enumerate(follow_headers, start=1):
        cell = ws_follow.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(FOLLOW_UP_MESSAGES, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_follow.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    email_headers = ["Тип письма", "Тема", "Текст", "Когда использовать"]
    for col, header in enumerate(email_headers, start=1):
        cell = ws_email.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(EMAIL_MESSAGES, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_email.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    process_headers = ["Блок", "Рекомендация"]
    for col, header in enumerate(process_headers, start=1):
        cell = ws_process.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(PROCESS_RECOMMENDATIONS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_process.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_social.freeze_panes = "A2"
    ws_follow.freeze_panes = "A2"
    ws_email.freeze_panes = "A2"
    ws_process.freeze_panes = "A2"

    social_widths = {"A": 24, "B": 16, "C": 90, "D": 32, "E": 28}
    follow_widths = {"A": 18, "B": 26, "C": 90}
    email_widths = {"A": 18, "B": 44, "C": 110, "D": 32}
    process_widths = {"A": 20, "B": 100}
    for column, width in social_widths.items():
        ws_social.column_dimensions[column].width = width
    for column, width in follow_widths.items():
        ws_follow.column_dimensions[column].width = width
    for column, width in email_widths.items():
        ws_email.column_dimensions[column].width = width
    for column, width in process_widths.items():
        ws_process.column_dimensions[column].width = width

    wb.save(MESSAGES_WORKBOOK_PATH)
    print(f"Создан файл '{MESSAGES_WORKBOOK_PATH.name}'.")


def build_sheet() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]

    ws_out = wb.create_sheet(title=OUTPUT_SHEET)
    for col_idx, header in enumerate(STANDARD_COLUMNS, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    kept_rows = 0
    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        header_map = normalize_header_map(ws)
        for row in range(2, ws.max_row + 1):
            practice = cell_value(ws, row, header_map, "ПРАКТИКУМ")
            city = cell_value(ws, row, header_map, "ГОРОД")
            fio = cell_value(ws, row, header_map, "ФИО")
            phone = cell_value(ws, row, header_map, "Номер телефона")
            tg = cell_value(ws, row, header_map, "Телеграм")
            responsible = cell_value(ws, row, header_map, "Ответственный")
            industry = cell_value(ws, row, header_map, "ИНДУСТРИЯ")
            company = cell_value(ws, row, header_map, "КОМПАНИЯ")
            site = cell_value(ws, row, header_map, "САЙТ")

            if not any([practice, city, fio, phone, tg, responsible, industry, company, site]):
                continue
            if should_skip_responsible(responsible):
                continue

            responsible_column = next(
                column for header, column in header_map.items() if "Ответственный" in header
            )
            ws.cell(row, responsible_column, value="Влад")

            added_info = ADDED_CONTACTS.get((sheet_name, row), default_added_info(phone, tg, site))
            add_row(
                ws_out,
                [
                    f"{sheet_name}!{row}",
                    practice,
                    city,
                    fio,
                    phone,
                    tg,
                    "Влад",
                    industry,
                    company,
                    site,
                ],
                added_info,
            )
            kept_rows += 1

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions
    autosize(ws_out)
    wb.save(WORKBOOK_PATH)
    print(f"Создан лист '{OUTPUT_SHEET}' с {kept_rows} строками.")


if __name__ == "__main__":
    build_sheet()
    build_messages_workbook()
