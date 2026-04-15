#!/usr/bin/env python3
"""
Generate consolidated contacts Excel and outreach message templates.
"""
import sys
sys.path.insert(0, '/home/ubuntu/.local/lib/python3.12/site-packages')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

INPUT_FILE = '/workspace/Сколково контакты.xlsx'
OUTPUT_CONTACTS = '/workspace/Контакты_для_Влада.xlsx'
OUTPUT_MESSAGES = '/workspace/Шаблоны_сообщений.xlsx'

EXCLUDE_NAMES = ['аня', 'анна', 'лена', 'елена', 'юля']

ADDED_INFO_FONT = Font(color='808080', italic=True)  # grey italic for added info
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

SHEETS_CONFIG = {
    'Екатеринбург': {'praktikum': 'A', 'fio': 'B', 'phone': 'C', 'telegram': 'D', 'responsible': 'E', 'industry': 'F', 'company': 'G', 'city': 'H', 'site': 'I'},
    'Казань': {'praktikum': 'A', 'fio': 'B', 'phone': 'C', 'telegram': 'D', 'responsible': 'E', 'industry': 'F', 'company': 'G', 'city': 'H', 'site': 'I'},
    'Нижний Новгород': {'praktikum': 'A', 'fio': 'B', 'phone': 'C', 'telegram': 'D', 'responsible': 'E', 'industry': 'F', 'company': 'G', 'city': 'H', 'site': 'I'},
    'Краснодарский край': {'praktikum': 'A', 'fio': 'C', 'phone': 'D', 'telegram': 'E', 'responsible': 'F', 'industry': 'G', 'company': 'H', 'city': 'B', 'site': 'I'},
}

# Enriched contact data from web searches
ENRICHED = {
    'Леонтьев Дмитрий': {
        'phone_add': '+7 (343) 283-07-37',
        'email': 'asp@altekproekt.ru',
        'site_add': 'altekproekt.ru',
        'notes': 'ООО «АЛЬТЕК СТРОЙ ПРОЕКТ», Екатеринбург, ул. Первомайская, 15, оф. 1201',
    },
    '?? Коваленко Светлана': {
        'phone_add': '+7 (343) 345-03-42',
        'email': 'expertural@expertural.com',
        'notes': 'Зам. директора «Эксперт-Урал», г. Екатеринбург, ул. Малышева, 105, оф. 619',
    },
    'Ярмышева Мария': {
        'phone_add': '+7 (343) 206-12-06',
        'site_add': 'yarmysheva.ru',
        'notes': 'Бренд ЯRMYSHEVA, Екатеринбург, пр-кт Ленина, 50Д, оф. 205-208',
    },
    'Вихарев Алексей': {
        'phone_add': '+7 (922) 130-44-33',
        'email': 'aviharev-vc@yandex.ru',
        'site_add': 'aviharev.ru',
        'notes': 'Депутат Екатеринбургской гор. Думы, адвокат, руководитель фракции «Единой России»',
    },
    'Зайченко Иван': {
        'telegram_add': '@ivanzaichenko',
        'notes': 'Основатель «Жизньмарт» (200+ магазинов), «Сушкоф и пицца». Instagram: @ivanzaychenko',
    },
    'Ларин Дмитрий': {
        'email': 'Dmitry.Larin@okkam.ru',
        'site_add': 'okkam.group/team/dmitrij_larin',
        'notes': 'Директор по трейд-маркетингу Okkam, 17 лет в аналитике для ритейла и FMCG. Москва, Пресненская наб. 6/2',
    },
    'Кузовков Василий': {
        'site_add': 'энергософт.рф',
        'notes': 'ООО ИК «Энергософт», Екатеринбург, пр. Решетникова, 22а, оф. 206. Выручка 867 млн руб.',
    },
    'Максимов Андрей': {
        'email': 'mr@tmr.rest',
        'site_add': 'shrmps-brgrs.ru, comunale.rest',
        'notes': 'Сеть «Креветки и бургеры», Comunale, Морская/10, Pumpula, Jackie. Для СМИ/партнерств: mr@tmr.rest',
    },
    'Мухаметшин Руслан': {
        'site_add': 'goodcom.rest',
        'notes': 'Холдинг Good Community: Breadway, The Barbara, Tribu, Encore Cafe, гастромолл «Главный». ул. Бориса Ельцина, 6',
    },
    'Калетин Андрей': {
        'phone_add': '+7 (343) 300-40-11',
        'email': 'info@ema.su',
        'site_add': 'emaholding.ru',
        'notes': 'Холдинг «ЭМА», Верх-Исетский б-р, 13, Екатеринбург. Экспорт медоборудования в 60+ стран',
    },
    'Маркова Екатерина': {
        'phone_add': '+7 (343) 305-19-84',
        'email': 'newbusiness@19agency84.com',
        'site_add': '19agency84.ru',
        'notes': 'Креативно-диджитал-агентство 19agency84, ул. Малышева, 71а, оф. 210. ~50-60 чел. в команде',
    },
    'Оглоблин Александр': {
        'phone_add': '+7 (343) 373-34-33',
        'email': 'ogloblin@list.ru',
        'notes': 'Развивает гастрономы «Елисейский» (4 точки), «Фабрика еды Гебо», франшиза «Бар здоровых привычек»',
    },
    'Бардок Алексей': {
        'site_add': 'initki.ru, get.initki.ru',
        'notes': 'Ген. директор ООО «Мотки и Бобины», онлайн-школа вязания и интернет-магазин iNitki',
    },
    'Абрамов Александр': {
        'notes': 'ООО «РЕСУРС», Набережные Челны. Оптовая торговля стройматериалами, производство, аренда',
    },
    'Измайлова Марина': {
        'site_add': 'marinaizmaylova.ru',
        'notes': 'Эксперт по созданию рентабельных пищевых производств под ключ, Казань',
    },
    'Шадриков Александр': {
        'email': 'eco@tatar.ru',
        'site_add': 'eco.tatarstan.ru',
        'notes': 'Министр экологии и природных ресурсов РТ с 2018 г. Адрес: ул. Павлюхина, 75, Казань',
    },
    'Яковлев Станислав': {
        'phone_add': '+7 (843) 265-55-88',
        'site_add': 'romangroup.ru',
        'notes': 'Президент Федерации скалолазания РТ. Председатель совета директоров ГК РоманГрупп (1000+ спортплощадок, ~50 парков)',
    },
    'Корнилов Антон': {
        'phone_add': '+7 (800) 600-17-25',
        'email': 'MyCornerMarketing@yandex.ru',
        'site_add': 'mycorner.ru',
        'notes': 'My Corner by Unistroy. Казань, ул. Николая Ершова, 76/1, пом. 109. Часть АО «Джи-групп»',
    },
    'Минуллина Талия': {
        'phone_add': '+7 (843) 570-40-01',
        'email': 'tida@tatar.ru',
        'site_add': 'tida.tatarstan.ru',
        'notes': 'Руководитель Агентства инвестиционного развития РТ с 2014 г. MBA, член Правительства РТ',
    },
    'Денисов Виталий': {
        'phone_add': '+7 (800) 700-51-40',
        'email': 'albina@imgevent.ru',
        'site_add': 'imgevent.ru',
        'notes': 'Сооснователь Imagine Group. Казань, ул. Островского, 87, оф. 502. Офисы в Казани и Москве, 55+ сотрудников',
    },
    'Гузь Роман': {
        'phone_add': '+7 (800) 700-51-40',
        'email': 'imgevent@mail.ru',
        'site_add': 'imgevent.ru',
        'notes': 'Сооснователь Imagine Group (совместно с Виталием Денисовым). Ивент-агентство, Казань',
    },
    'Боткин Евгений': {
        'phone_add': '+7 (831) 412-32-17',
        'email': 'vita@kis.ru',
        'site_add': 'vita-print.com',
        'notes': 'Ген. директор НПП «Вита-Принт» с 2007 г. Н. Новгород, ул. Бекетова, 13, стр. 3. Выручка 387 млн (2024), 94 сотр.',
    },
    'Козлов Никита': {
        'phone_add': '+7 (83130) 6-77-77 доб. 153',
        'email': 'zakaz@globaltest.ru',
        'site_add': 'globaltest.ru',
        'notes': 'Учредитель ООО «ГЛОБАЛТЕСТ». Саров, ул. Павлика Морозова, 6. Офис в Москве: Долгопрудненское ш., 3',
    },
    'Суфиянова Евгения': {
        'email': 'info@gastreet.com',
        'notes': 'Покидает GASTREET, запустила HORECA HUB. Контакт пресс-службы GASTREET: pr@gastreet.com, тел. 8 (800) 700-93-20',
    },
    'Трукшин Вадим': {
        'phone_add': '8 (800) 100-78-62',
        'email': 'mice@mantera-group.com',
        'site_add': 'mantera.ru',
        'notes': 'Ген. директор MANTERA с 2023 г. Инвестпортфель 100+ млрд руб., 7.1 млн туристов/год. Сириус, Континентальный пр-кт, 6',
    },
    'Перов Павел': {
        'phone_add': '+7 (800) 550-20-20',
        'email': 'pr@kpresort.ru',
        'site_add': 'krasnayapolyanaresort.ru',
        'notes': 'Ген. директор Курорта Красная Поляна с 06.2024. Выпускник Сколково. Sales Office Москва: Пресненская наб. 12, стр 2',
    },
    'Барачина Надежда': {
        'phone_add': '+7 (861) 205-09-84',
        'email': 'info@alias-group.ru',
        'site_add': 'alias-group.ru',
        'notes': 'Ген. директор Alias Group с 2021. Краснодар, ул. Комсомольская, 15, оф. 126. 118 сотрудников',
    },
    'Манькова Инна': {
        'phone_add': '8 (800) 700-1-800',
        'email': 'bsfc@bsfc.com',
        'site_add': 'bsfc.com',
        'notes': 'Директор ООО «ЧФК — Недвижимость» (Черноморская финансовая компания). Краснодар, Красная ул., 108',
    },
    'Прядка Алексей': {
        'email': 'o@modern-vitamins.ru',
        'notes': 'ООО «Ванвин». Выручка 205 млн (2024). Витамины, БАД, спортивное питание',
    },
    'Рунец Роман': {
        'site_add': 'itr2050.ru',
        'notes': 'Ген. директор ООО «ИТР» с 02.2025. Краснодар, ул. Октябрьская, 59',
    },
    'Мищенко Игорь': {
        'notes': 'ИП, деятельность агентств недвижимости, строительство. Краснодарский край',
    },
}


def should_exclude(responsible_val):
    if not responsible_val:
        return False
    resp_lower = str(responsible_val).lower()
    return any(name in resp_lower for name in EXCLUDE_NAMES)


def get_cell_val(ws, col, row):
    val = ws[f'{col}{row}'].value
    return val if val is not None else ''


def create_contacts_excel():
    wb_in = openpyxl.load_workbook(INPUT_FILE)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Контакты для Влада'

    headers = ['№', 'Регион', 'Город', 'ФИО', 'Телефон', 'Email', 'Телеграм',
               'Ответственный за контакт', 'Индустрия', 'Компания', 'Сайт', 'Примечания']

    for col_idx, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    ws_out.row_dimensions[1].height = 30

    row_num = 2
    contact_num = 1

    for sheet_name, cols in SHEETS_CONFIG.items():
        ws_in = wb_in[sheet_name]

        for src_row in range(2, ws_in.max_row + 1):
            fio = get_cell_val(ws_in, cols['fio'], src_row)
            if not fio or str(fio).strip() == '':
                continue

            resp = get_cell_val(ws_in, cols['responsible'], src_row)
            if should_exclude(resp):
                continue

            fio_str = str(fio).strip()
            phone = str(get_cell_val(ws_in, cols['phone'], src_row)).strip()
            telegram = str(get_cell_val(ws_in, cols['telegram'], src_row)).strip()
            industry = str(get_cell_val(ws_in, cols['industry'], src_row)).strip()
            company = str(get_cell_val(ws_in, cols['company'], src_row)).strip()
            city = str(get_cell_val(ws_in, cols['city'], src_row)).strip()
            site = str(get_cell_val(ws_in, cols['site'], src_row)).strip()

            enriched = ENRICHED.get(fio_str, {})

            if enriched.get('phone_add') and phone:
                phone = f"{phone}; {enriched['phone_add']}"
            elif enriched.get('phone_add'):
                phone = enriched['phone_add']

            email = enriched.get('email', '')

            if enriched.get('telegram_add') and telegram:
                telegram = f"{telegram}; {enriched['telegram_add']}"
            elif enriched.get('telegram_add'):
                telegram = enriched['telegram_add']

            if enriched.get('site_add') and site:
                site = f"{site}; {enriched['site_add']}"
            elif enriched.get('site_add'):
                site = enriched['site_add']

            notes = enriched.get('notes', '')

            row_data = [contact_num, sheet_name, city, fio_str, phone, email, telegram,
                        'Влад', industry, company, site, notes]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws_out.cell(row=row_num, column=col_idx, value=val if val else '')
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = THIN_BORDER

                if col_idx in (5, 6, 7, 11, 12) and val and fio_str in ENRICHED:
                    orig_vals = {
                        5: str(get_cell_val(ws_in, cols['phone'], src_row)).strip(),
                        6: '',
                        7: str(get_cell_val(ws_in, cols['telegram'], src_row)).strip(),
                        11: str(get_cell_val(ws_in, cols['site'], src_row)).strip(),
                        12: '',
                    }
                    if str(val) != orig_vals.get(col_idx, ''):
                        cell.font = ADDED_INFO_FONT

            if row_num % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws_out.cell(row=row_num, column=col_idx).fill = PatternFill(
                        start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

            row_num += 1
            contact_num += 1

    col_widths = [5, 20, 18, 25, 22, 28, 25, 18, 30, 50, 35, 50]
    for i, w in enumerate(col_widths, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    ws_out.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_num - 1}'
    ws_out.freeze_panes = 'A2'

    wb_out.save(OUTPUT_CONTACTS)
    print(f'Saved contacts: {OUTPUT_CONTACTS} ({contact_num - 1} contacts)')


def create_messages_excel():
    wb = openpyxl.Workbook()

    # Sheet 1: Social media messages (Telegram/VK)
    ws1 = wb.active
    ws1.title = 'Соцсети — Первое касание'

    messages_social = [
        {
            'name': 'Вариант 1 — Мягкий заход (2 сообщения)',
            'msg1': (
                'Добрый день, [Имя]! Меня зовут Влад, я из фонда «БольшеЧемМожешь» — '
                'мы развиваем инклюзивный спорт и помогаем людям с особенностями здоровья '
                'участвовать в забегах и марафонах.\n\n'
                'Пишу вам, потому что 20 июня на кампусе Школы управления СКОЛКОВО пройдет '
                'благотворительный забег «Сколковская миля». Хотел бы обсудить возможное участие '
                'вашей компании в качестве партнера. Могу коротко рассказать о форматах — '
                'удобно будет здесь или лучше по почте?'
            ),
            'msg2': '',
            'comment': (
                'Один компактный текст, сразу понятно кто пишет и зачем. '
                'Вопрос в конце снижает барьер ответа.'
            ),
        },
        {
            'name': 'Вариант 2 — Два коротких сообщения',
            'msg1': (
                'Добрый день, [Имя]! Это Влад из благотворительного фонда «БольшеЧемМожешь». '
                'Мы помогаем людям с ДЦП, аутизмом и другими особенностями здоровья заниматься спортом '
                'и преодолевать марафонские дистанции.'
            ),
            'msg2': (
                '20 июня проводим забег «Сколковская миля» совместно со СКОЛКОВО — '
                'на крыше кампуса, 300+ участников, медиаохват 20 000+. '
                'Ищем партнеров — есть пакеты от 100 000 ₽. '
                'Могу скинуть презентацию, если интересно?'
            ),
            'comment': (
                'Два сообщения: первое знакомит, второе — конкретное предложение. '
                'Работает лучше для мессенджеров, где длинные тексты пугают.'
            ),
        },
        {
            'name': 'Вариант 3 — Через ценность для бизнеса',
            'msg1': (
                'Добрый день, [Имя]! Меня зовут Влад, я представляю фонд «БольшеЧемМожешь».\n\n'
                'Хочу предложить интересную возможность для [название компании]: '
                '20 июня Школа управления СКОЛКОВО и наш фонд проводят благотворительный забег '
                '«Сколковская миля» — 300+ участников (предприниматели, выпускники СКОЛКОВО, '
                'топ-менеджеры). Партнерство даёт узнаваемость бренда в деловом сообществе, '
                'интеграцию продукта на площадке и PR-охват.\n\n'
                'Удобно будет коротко обсудить по телефону или в переписке?'
            ),
            'msg2': '',
            'comment': (
                'Акцент на выгодах для бизнеса: аудитория, узнаваемость, PR. '
                'Подходит для крупных компаний.'
            ),
        },
        {
            'name': 'Вариант 4 — Персональный (для знакомых по Практикуму)',
            'msg1': (
                '[Имя], привет! Это Влад из фонда «БольшеЧемМожешь» — '
                'мы пересекались на Практикуме в Сколково.\n\n'
                'Хотел рассказать про наш ближайший проект: 20 июня проводим забег '
                '«Сколковская миля» на крыше кампуса Школы. Ищем партнёров, которые разделяют '
                'ценности инклюзии и активного образа жизни.\n\n'
                'Есть несколько минут обсудить? Могу скинуть короткую презентацию.'
            ),
            'msg2': '',
            'comment': (
                'Персональный тон, отсылка к общему опыту (Практикум). '
                'Для людей, которые проходили обучение в Сколково.'
            ),
        },
    ]

    headers1 = ['Вариант', 'Сообщение 1', 'Сообщение 2 (если есть)', 'Комментарий']
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    for i, msg in enumerate(messages_social, 2):
        ws1.cell(row=i, column=1, value=msg['name']).alignment = Alignment(vertical='top', wrap_text=True)
        ws1.cell(row=i, column=2, value=msg['msg1']).alignment = Alignment(vertical='top', wrap_text=True)
        ws1.cell(row=i, column=3, value=msg['msg2']).alignment = Alignment(vertical='top', wrap_text=True)
        ws1.cell(row=i, column=4, value=msg['comment']).alignment = Alignment(vertical='top', wrap_text=True)
        for c in range(1, 5):
            ws1.cell(row=i, column=c).border = THIN_BORDER
        ws1.row_dimensions[i].height = 120

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 60
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 40

    # Sheet 2: Email templates
    ws2 = wb.create_sheet('Email — Деловое письмо')

    email_templates = [
        {
            'name': 'Вариант 1 — Деловой',
            'subject': 'Партнёрство в забеге «Сколковская миля» — 20 июня, СКОЛКОВО',
            'body': (
                'Добрый день, [Имя]!\n\n'
                'Меня зовут Владислав, я фандрайзер Благотворительного Фонда «БольшеЧемМожешь» '
                '(morethanable.ru).\n\n'
                'Наш фонд с 2016 года развивает инклюзивный спорт: помогает людям с ДЦП, '
                'аутизмом и другими особенностями здоровья заниматься спортом и финишировать '
                'марафоны. Более 300 подопечных, 700 волонтёров, 48 беговых стартов в год.\n\n'
                '20 июня 2026 года совместно с Московской школой управления СКОЛКОВО '
                'мы проводим благотворительный забег «Сколковская миля» на крыше кампуса '
                'Школы. Это традиционное событие, открывающее День выпускника СКОЛКОВО:\n\n'
                '• 300+ участников — предприниматели, выпускники Школы, топ-менеджеры\n'
                '• Медиаохват — 20 000+\n'
                '• Дистанция — 2006 метров (в честь года основания Школы)\n\n'
                'Мы ищем партнёров, которые хотели бы поддержать мероприятие и получить '
                'присутствие бренда на площадке. Партнёрские пакеты от 100 000 ₽ — с размещением '
                'логотипа, упоминанием ведущим, интеграцией продукта/сервиса и PR-поддержкой.\n\n'
                'Прикладываю презентацию с подробностями. Буду рад обсудить удобный формат '
                'участия для вашей компании.\n\n'
                'С уважением,\n'
                'Владислав\n'
                'Фандрайзер БФ «БольшеЧемМожешь»\n'
                'morethanable.ru'
            ),
            'comment': 'Стандартный деловой формат. Подходит для холодных контактов.',
        },
        {
            'name': 'Вариант 2 — Короткий',
            'subject': '«Сколковская миля» 20 июня — партнёрство',
            'body': (
                'Добрый день, [Имя]!\n\n'
                'Владислав, фонд «БольшеЧемМожешь».\n\n'
                '20 июня на кампусе СКОЛКОВО проводим благотворительный забег — '
                '300+ участников из делового сообщества Школы, медиаохват 20 000+.\n\n'
                'Партнёрские пакеты — от 100 тыс. ₽ (брендинг на площадке, '
                'упоминания ведущим, PR в каналах Школы и фонда).\n\n'
                'Презентация во вложении. Готов обсудить в удобном формате.\n\n'
                'С уважением,\n'
                'Владислав\n'
                'Фандрайзер БФ «БольшеЧемМожешь»\n'
                'morethanable.ru'
            ),
            'comment': 'Краткий вариант. Для занятых людей, которые не будут читать длинные письма.',
        },
        {
            'name': 'Вариант 3 — Эмоциональный',
            'subject': 'Забег, который меняет жизни — ищем партнёров',
            'body': (
                'Добрый день, [Имя]!\n\n'
                'Меня зовут Владислав, я из фонда «БольшеЧемМожешь».\n\n'
                'С нами люди, которые не могут сделать даже шаг, финишируют марафоны. '
                'С 2016 года мы помогаем людям с ДЦП, аутизмом и другими особенностями '
                'здоровья — через спорт, через включение в сообщество, через веру в то, '
                'что каждый может больше.\n\n'
                '20 июня мы проводим «Сколковскую милю» на крыше кампуса Школы управления '
                'СКОЛКОВО. 300+ участников пробегут 2006 метров в поддержку наших подопечных.\n\n'
                'Мы ищем партнёров, для которых социальная ответственность — '
                'не строчка в отчёте, а реальное действие. Есть несколько пакетов участия '
                '(от 100 000 ₽) с интеграцией бренда на площадке.\n\n'
                'Презентация — во вложении. Буду рад обсудить, как это может быть '
                'интересно именно вашей компании.\n\n'
                'С уважением,\n'
                'Владислав\n'
                'Фандрайзер БФ «БольшеЧемМожешь»\n'
                'morethanable.ru'
            ),
            'comment': 'Эмоциональный тон с акцентом на миссию. Для людей, ценящих социальный импакт.',
        },
    ]

    headers2 = ['Вариант', 'Тема письма', 'Текст письма', 'Комментарий']
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    for i, tpl in enumerate(email_templates, 2):
        ws2.cell(row=i, column=1, value=tpl['name']).alignment = Alignment(vertical='top', wrap_text=True)
        ws2.cell(row=i, column=2, value=tpl['subject']).alignment = Alignment(vertical='top', wrap_text=True)
        ws2.cell(row=i, column=3, value=tpl['body']).alignment = Alignment(vertical='top', wrap_text=True)
        ws2.cell(row=i, column=4, value=tpl['comment']).alignment = Alignment(vertical='top', wrap_text=True)
        for c in range(1, 5):
            ws2.cell(row=i, column=c).border = THIN_BORDER
        ws2.row_dimensions[i].height = 250

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 70
    ws2.column_dimensions['D'].width = 40

    # Sheet 3: Follow-up messages
    ws3 = wb.create_sheet('Фоллоу-ап')

    followups = [
        {
            'name': 'Через 2-3 дня (нет ответа)',
            'social': (
                '[Имя], добрый день! Писал вам пару дней назад по поводу забега «Сколковская миля» '
                '20 июня в Сколково. Понимаю, что может быть загруженное время — '
                'просто хотел убедиться, что сообщение дошло. '
                'Если тема в целом не актуальна — скажите, не буду отвлекать.'
            ),
            'email_subject': 'Re: Партнёрство в забеге «Сколковская миля»',
            'email_body': (
                '[Имя], добрый день!\n\n'
                'Писал вам ранее по поводу партнёрства в забеге «Сколковская миля» 20 июня. '
                'Хотел уточнить — удалось ли посмотреть презентацию?\n\n'
                'Если тема интересна — с удовольствием расскажу подробнее. '
                'Если сейчас не актуально — тоже пойму, не буду отвлекать.\n\n'
                'С уважением, Владислав'
            ),
            'comment': 'Мягкий фоллоу-ап. Даёт выход из разговора, снижает давление.',
        },
        {
            'name': 'Через 5-7 дней (второй фоллоу-ап)',
            'social': (
                '[Имя], здравствуйте! Возвращаюсь к теме забега «Сколковская миля». '
                'До события остаётся [X] недель, и мы формируем финальный пул партнёров. '
                'Если вам интересно — готов обсудить любой удобный формат участия. '
                'Буду рад обратной связи!'
            ),
            'email_subject': 'Re: Партнёрство в забеге «Сколковская миля» — формируем партнёров',
            'email_body': (
                '[Имя], добрый день!\n\n'
                'Возвращаюсь к теме партнёрства в «Сколковской миле» 20 июня.\n\n'
                'Сейчас формируем финальный список партнёров мероприятия. '
                'Если вам была бы интересна короткая встреча или звонок на 10 минут — '
                'с удовольствием расскажу, как партнёрство может быть полезно '
                'именно вашей компании.\n\n'
                'С уважением, Владислав\n'
                'Фандрайзер БФ «БольшеЧемМожешь»'
            ),
            'comment': 'Создаёт лёгкую срочность (формирование пула). Предлагает короткий звонок.',
        },
    ]

    headers3 = ['Этап', 'Текст для соцсетей', 'Тема email', 'Текст email', 'Комментарий']
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    for i, fu in enumerate(followups, 2):
        ws3.cell(row=i, column=1, value=fu['name']).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(row=i, column=2, value=fu['social']).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(row=i, column=3, value=fu['email_subject']).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(row=i, column=4, value=fu['email_body']).alignment = Alignment(vertical='top', wrap_text=True)
        ws3.cell(row=i, column=5, value=fu['comment']).alignment = Alignment(vertical='top', wrap_text=True)
        for c in range(1, 6):
            ws3.cell(row=i, column=c).border = THIN_BORDER
        ws3.row_dimensions[i].height = 150

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 55
    ws3.column_dimensions['C'].width = 35
    ws3.column_dimensions['D'].width = 55
    ws3.column_dimensions['E'].width = 35

    # Sheet 4: Tips
    ws4 = wb.create_sheet('Рекомендации')

    tips = [
        ['Рекомендации по процессу рассылки', ''],
        ['1. Приоритизация контактов', 'Начинайте с тех, у кого есть Telegram — это самый быстрый канал. '
         'Далее — те, у кого есть email. В последнюю очередь — те, у кого только телефон (звонок).'],
        ['2. Время отправки', 'Соцсети: Вт-Чт, 10:00-12:00 или 15:00-17:00. '
         'Email: Вт-Чт, 9:00-11:00. Понедельник и пятница дают худший отклик.'],
        ['3. Персонализация', 'Всегда подставляйте имя и название компании в шаблон. '
         'Если знаете что-то о человеке (проект, достижение) — упомяните в первых строках.'],
        ['4. Темп рассылки', 'Не более 15-20 сообщений в день в мессенджерах (во избежание блокировки аккаунта). '
         'Email — до 30-40 в день, лучше через mail merge.'],
        ['5. CRM / трекинг', 'Ведите таблицу отправок: дата, канал, статус (отправлено/прочитано/ответ/отказ). '
         'Это поможет не потерять контакты и планировать фоллоу-апы.'],
        ['6. Дублирование каналов', 'Сначала пишите в мессенджер. Если нет ответа 2 дня — дублируйте на email. '
         'Это повышает шансы на ответ в 2-3 раза.'],
        ['7. Сегментация текста', 'Для HoReCa/ритейла делайте акцент на аудиторию и PR-охват. '
         'Для строительных/промышленных компаний — на HR-ценность и корпоративный тимбилдинг. '
         'Для госструктур — на социальную ответственность и имидж.'],
        ['8. Подготовка к разговору', 'Имейте под рукой: презентацию PDF, 2-3 фото/видео с прошлых забегов, '
         'ссылку на сайт фонда. Люди просят доп. материалы в 70% случаев.'],
        ['9. Отказы — это нормально', 'Конверсия холодной базы в партнёрство: 3-7%. '
         'Из 75 контактов реалистично привлечь 3-5 партнёров. '
         'Не воспринимайте отказы лично.'],
        ['10. Выпускники Сколково', 'Контакты из Практикумов Сколково — ваш лучший актив. '
         'У них уже есть связь со Школой, и забег на крыше кампуса для них — '
         'дополнительный аргумент.'],
    ]

    for i, (title, desc) in enumerate(tips, 1):
        cell_a = ws4.cell(row=i, column=1, value=title)
        cell_b = ws4.cell(row=i, column=2, value=desc)
        cell_a.alignment = Alignment(vertical='top', wrap_text=True)
        cell_b.alignment = Alignment(vertical='top', wrap_text=True)
        if i == 1:
            cell_a.font = Font(bold=True, size=14)
        else:
            cell_a.font = Font(bold=True, size=11)
        ws4.row_dimensions[i].height = 50

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 80

    wb.save(OUTPUT_MESSAGES)
    print(f'Saved messages: {OUTPUT_MESSAGES}')


if __name__ == '__main__':
    create_contacts_excel()
    create_messages_excel()
    print('Done!')
